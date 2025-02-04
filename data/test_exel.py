from openpyxl import load_workbook

# Charger le fichier Excel
workbook = load_workbook("nom_du_fichier_modifié.xlsx")

# Sélectionner une feuille
sheet = workbook["F1"]

# Lire une cellule
valeur = sheet["A2"].value


sheet["A2"].value = 1


# Sauvegarder les modifications
workbook.save("nom_du_fichier_modifié.xlsx")


fichier = load_workbook("nom_du_fichier_modifié.xlsx")
page = fichier["F1"]

i = 1
v = sheet["A"+str(i)].value
while(v!=None):
    i+=1
    v = sheet["A"+str(i)].value
    


page["A"+str(i)].value = i
fichier.save("nom_du_fichier_modifié.xlsx")




from openpyxl import Workbook

# Créer un nouveau classeur
workbook = Workbook()

# Sélectionner la feuille active (créée par défaut)
sheet = workbook.active
sheet.title = "Feuiiiiille1"

# Ajouter des données dans des cellules spécifiques
sheet["A1"] = "Nom"
sheet["B1"] = "Âge"
sheet["C1"] = "Ville"
sheet["A2"] = "Alice"
sheet["B2"] = 25
sheet["C2"] = "Paris aux chiottes"
sheet["A3"] = "Bob"
sheet["B3"] = 30
sheet["C3"] = "Lyon"

# Enregistrer le fichier
workbook.save("lalalala.xlsx")


